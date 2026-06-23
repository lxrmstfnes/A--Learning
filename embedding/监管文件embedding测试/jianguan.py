#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
监管文件 Embedding 入库脚本
==========================

功能概述:
    1. 读取「监管文件」目录下的 PDF / DOC 监管文档（兼容 Win 转 Mac 后的乱码文件名）
    2. 借助目录 Excel 还原文件中文名称
    3. 调用阿里云百炼 Qwen3-Embedding（text-embedding-v4）生成向量
    4. 按「章 / 节 / 条」结构化切分后，使用 FAISS 本地持久化向量索引

用法:
    python jianguan.py                  # 构建索引
    python jianguan.py --search "数据安全"  # 语义检索
    python jianguan.py --rebuild        # 强制重建索引

依赖:
    pip install -r requirements.txt

API Key:
    优先读取环境变量 DASHSCOPE_API_KEY；若未设置，则从 ~/.zshrc 解析
    DASHSCOPE_API_KEY 或 OPENAI_API_KEY。
"""

from __future__ import annotations

import argparse
import json
import os
import pickle
import re
import subprocess
import sys
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import faiss
import numpy as np
import openpyxl
from openai import OpenAI

try:
    import fitz  # pymupdf
except ImportError as exc:  # pragma: no cover - 依赖缺失时给出明确提示
    raise ImportError("请先安装 pymupdf: pip install pymupdf") from exc


# =============================================================================
# 路径与模型配置
# =============================================================================

# 脚本所在目录
BASE_DIR = Path(__file__).resolve().parent

# 监管文件原始目录
DOC_DIR = BASE_DIR / "监管文件"

# FAISS 索引与元数据输出目录
INDEX_DIR = BASE_DIR / "faiss_index"

# FAISS 向量索引文件
FAISS_INDEX_FILE = INDEX_DIR / "jianguan.index"

# 文本块元数据（与向量一一对应）
METADATA_FILE = INDEX_DIR / "metadata.pkl"

# 索引配置信息（便于排查问题）
CONFIG_FILE = INDEX_DIR / "config.json"

# 百炼 Embedding 模型（Qwen3-Embedding 系列）
EMBEDDING_MODEL = "text-embedding-v4"

# 向量维度（text-embedding-v4 默认 1024，可按需调整）
EMBEDDING_DIM = 1024

# 百炼 OpenAI 兼容接口地址
DASHSCOPE_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"

# 单次 API 最多提交的文本条数（百炼限制为 10）
EMBED_BATCH_SIZE = 10

# 单条法规内容过长时的二次切分阈值（字符数）
MAX_ARTICLE_LEN = 1200

# 超长条款二次切分时的重叠长度
CHUNK_OVERLAP = 120

# 中文数字（用于匹配「第X章/条/点」及「X、」等结构）
CN_NUM = r"[一二三四五六七八九十百零\d]+"

# zshrc 中可能存放 API Key 的变量名（按优先级排列）
API_KEY_ENV_NAMES = ("DASHSCOPE_API_KEY", "OPENAI_API_KEY")

# 目录 Excel 文件名特征：内容以「编号」开头（乱码名也可识别）
CATALOG_XLSX_HEADER = "编号"

# 摘要 Excel 文件名特征：首列表头为「文件名」
SUMMARY_XLSX_HEADER = "文件名"

# 当目录 Excel 缺失时使用的内置映射（编号 / 前缀 / 年月 / 中文文件名）
FALLBACK_CATALOG: List[Tuple[int, str, str, str]] = [
    (1, "1", "2018-04", "关于规范金融机构资产管理业务的指导意见"),
    (2, "1", "2018-07", "关于进一步明确规范金融机构资产管理业务指导意见有关事项的通知"),
    (3, "N", "2018-09", "商业银行理财业务监督管理办法"),
    (4, "N", "2018-12", "商业银行理财子公司管理办法"),
    (5, "N", "2019-11", "商业银行理财子公司净资本管理办法（试行）"),
    (6, "N", "2021-05", "理财公司理财产品销售管理暂行办法"),
    (7, "N", "2021-12", "理财公司理财产品流动性风险管理办法"),
    (8, "N", "2022-09", "理财公司内部控制管理办法"),
    (9, "X", "2018-05", "银行业金融机构数据治理指引"),
    (10, "X", "2020-07", "标准化债权类资产认定规则"),
    (11, "X", "2021-06", "关于规范现金管理类理财产品管理有关事项的通知"),
    (12, "X", "2021-06", "银行保险机构公司治理准则"),
    (13, "X", "2021-09", "关于开展养老理财产品试点的通知"),
    (14, "X", "2021-11", "银行保险机构关联交易管理办法"),
    (15, "X", "2022-02", "中国银保监会办公厅关于扩大养老理财产品试点范围的通知"),
    (16, "X", "2022-11", "关于印发商业银行和理财公司个人养老金业务管理暂行办法的通知"),
    (17, "X", "2024-12", "关于公司治理监管规定与公司法衔接有关事项的通知"),
    (18, "X", "2024-12", "金融机构合规管理办法"),
    (19, "X", "2024-12", "银行保险机构数据安全管理办法"),
    (20, "X", "2025-03", "国家金融监督管理总局关于印发《商业银行代理销售业务管理办法》的通知"),
    (21, "X", "2025-04", "银行业金融机构董事（理事）和高级管理人员任职资格管理办法"),
    (22, "X", "2025-07", "金融机构产品适当性管理办法"),
]


# =============================================================================
# 数据结构
# =============================================================================


@dataclass
class CatalogEntry:
    """文件目录中的一条记录。"""

    doc_id: int
    category_prefix: str
    publish_date: datetime
    display_name: str


@dataclass
class ArticleChunk:
    """按法规结构切分后的单个片段（章/节/条级别）。"""

    text: str
    chapter: str = ""
    section: str = ""
    article: str = ""


@dataclass
class TextChunk:
    """待向量化的文本块及其溯源信息。"""

    chunk_id: int
    text: str
    source_type: str
    raw_filename: str
    display_name: str
    chapter: str = ""
    section: str = ""
    article: str = ""
    summary: str = ""


# =============================================================================
# API Key 加载
# =============================================================================


def load_api_key() -> str:
    """
    获取百炼 API Key。

    读取顺序:
        1. 环境变量 DASHSCOPE_API_KEY
        2. 环境变量 OPENAI_API_KEY
        3. ~/.zshrc 中的 export 语句
    """
    for name in API_KEY_ENV_NAMES:
        value = os.getenv(name, "").strip()
        if value:
            return value

    zshrc_path = Path.home() / ".zshrc"
    if not zshrc_path.exists():
        raise EnvironmentError(
            "未找到 API Key。请在 ~/.zshrc 中设置 export DASHSCOPE_API_KEY=xxx"
        )

    content = zshrc_path.read_text(encoding="utf-8", errors="ignore")
    pattern = re.compile(
        r"export\s+(DASHSCOPE_API_KEY|OPENAI_API_KEY)\s*=\s*['\"]?([^'\"\n#]+)['\"]?"
    )
    found: Dict[str, str] = {}
    for env_name, env_value in pattern.findall(content):
        found[env_name] = env_value.strip()

    for name in API_KEY_ENV_NAMES:
        if name in found and found[name]:
            return found[name]

    raise EnvironmentError(
        "在环境变量和 ~/.zshrc 中均未找到 DASHSCOPE_API_KEY 或 OPENAI_API_KEY。"
    )


def create_embedding_client(api_key: str) -> OpenAI:
    """创建百炼 OpenAI 兼容客户端。"""
    return OpenAI(api_key=api_key, base_url=DASHSCOPE_BASE_URL)


# =============================================================================
# 文件名修复与目录映射
# =============================================================================


def excel_serial_to_datetime(serial: int) -> datetime:
    """将 Excel 日期序列号转换为 datetime。"""
    return datetime(1899, 12, 30) + timedelta(days=int(serial))


def find_xlsx_by_header(doc_dir: Path, header_value: str) -> Optional[Path]:
    """
    在目录中查找指定表头的 xlsx 文件。

    说明:
        Win 转 Mac 后 xlsx 文件名可能乱码，因此通过表头内容识别文件类型。
    """
    for path in sorted(doc_dir.glob("*.xlsx")):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        first_cell = worksheet.cell(row=1, column=1).value
        workbook.close()
        if first_cell == header_value:
            return path
    return None


def load_catalog(doc_dir: Path) -> List[CatalogEntry]:
    """读取「文件目录.xlsx」，构建编号 -> 中文文件名映射；缺失时使用内置目录。"""
    catalog_path = find_xlsx_by_header(doc_dir, CATALOG_XLSX_HEADER)
    entries: List[CatalogEntry] = []

    if catalog_path is not None:
        workbook = openpyxl.load_workbook(catalog_path, read_only=True, data_only=True)
        worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            doc_id, _, publish_serial, display_name = row[0], row[1], row[2], row[3]
            if not isinstance(doc_id, int) or not display_name:
                continue

            if doc_id <= 2:
                prefix = "1"
            elif doc_id <= 8:
                prefix = "N"
            else:
                prefix = "X"

            publish_date = (
                excel_serial_to_datetime(publish_serial)
                if publish_serial
                else datetime(1900, 1, 1)
            )
            entries.append(
                CatalogEntry(
                    doc_id=doc_id,
                    category_prefix=prefix,
                    publish_date=publish_date,
                    display_name=str(display_name).strip(),
                )
            )

        workbook.close()
        return entries

    # Excel 不存在时，使用内置目录兜底
    for doc_id, prefix, year_month, display_name in FALLBACK_CATALOG:
        year, month = year_month.split("-")
        entries.append(
            CatalogEntry(
                doc_id=doc_id,
                category_prefix=prefix,
                publish_date=datetime(int(year), int(month), 1),
                display_name=display_name,
            )
        )
    return entries


def parse_file_year_month(filename: str) -> Tuple[str, str]:
    """
    从乱码文件名中解析分类前缀与年月。

    示例:
        1-2018...4...pdf  -> ("1", "2018-04")
        N-2021...5...doc  -> ("N", "2021-05")
    """
    prefix = filename[0] if filename and filename[0] in {"1", "N", "X"} else "?"
    match = re.search(r"(20\d{2}).*?(\d{1,2})", filename)
    if not match:
        return prefix, ""
    year = match.group(1)
    month = int(match.group(2))
    return prefix, f"{year}-{month:02d}"


def score_name_by_text(text: str, display_name: str) -> int:
    """用正文关键词与目录名称的重合度进行消歧。"""
    if not text or not display_name:
        return 0

    keywords = re.findall(r"[\u4e00-\u9fff]{2,}", display_name)
    sample = text[:3000]
    return sum(sample.count(keyword) for keyword in keywords)


def match_file_to_catalog(
    filename: str,
    preview_text: str,
    catalog: Sequence[CatalogEntry],
) -> CatalogEntry:
    """
    将磁盘上的乱码文件名映射到目录中的中文名称。

    匹配策略:
        1. 先按「分类前缀 + 年月」筛选候选
        2. 若候选唯一，直接返回
        3. 若候选多个，用正文与文件名的关键词重合度消歧
        4. 若仍无法匹配，返回占位名称
    """
    prefix, year_month = parse_file_year_month(filename)
    candidates = [
        item
        for item in catalog
        if item.category_prefix == prefix
        and item.publish_date.strftime("%Y-%m") == year_month
    ]

    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        ranked = sorted(
            candidates,
            key=lambda item: score_name_by_text(preview_text, item.display_name),
            reverse=True,
        )
        if ranked[0] and score_name_by_text(preview_text, ranked[0].display_name) > 0:
            return ranked[0]
        return ranked[0]

    # 兜底：只按前缀返回第一条，避免流程中断
    prefix_candidates = [item for item in catalog if item.category_prefix == prefix]
    if prefix_candidates:
        return prefix_candidates[0]

    return CatalogEntry(
        doc_id=-1,
        category_prefix=prefix,
        publish_date=datetime(1900, 1, 1),
        display_name=filename,
    )


# =============================================================================
# 文档读取
# =============================================================================


def extract_pdf_text(file_path: Path) -> str:
    """使用 PyMuPDF 提取 PDF 文本。"""
    texts: List[str] = []
    with fitz.open(file_path) as document:
        for page in document:
            page_text = page.get_text("text").strip()
            if page_text:
                texts.append(page_text)
    return "\n".join(texts)


def extract_doc_text(file_path: Path) -> str:
    """
    提取 legacy .doc 文本。

    macOS 优先调用系统自带的 textutil；失败时给出友好提示。
    """
    suffix = file_path.suffix.lower()
    if suffix == ".docx":
        try:
            import docx  # type: ignore
        except ImportError as exc:
            raise ImportError("读取 .docx 需要 python-docx，请先安装。") from exc
        document = docx.Document(str(file_path))
        return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text)

    # .doc 旧格式
    result = subprocess.run(
        ["textutil", "-convert", "txt", "-stdout", str(file_path)],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode == 0 and result.stdout.strip():
        return result.stdout.strip()

    raise RuntimeError(
        f"无法解析文档: {file_path.name}。"
        "如在非 macOS 环境，请将 .doc 转为 .docx 或 .pdf 后重试。"
    )


def extract_document_text(file_path: Path) -> str:
    """根据扩展名选择对应的文本抽取方式。"""
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf_text(file_path)
    if suffix in {".doc", ".docx"}:
        return extract_doc_text(file_path)
    raise ValueError(f"暂不支持的文件类型: {file_path.name}")


def normalize_text(text: str) -> str:
    """清理多余空白，便于分块与向量化。"""
    text = text.replace("\u2003", " ").replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def split_long_article(text: str, max_len: int = MAX_ARTICLE_LEN) -> List[str]:
    """
    对超长单条法规内容做二次切分（保留同一条款编号）。

    参数:
        text: 单条法规正文
        max_len: 允许的最大字符数
    """
    if len(text) <= max_len:
        return [text]

    pieces: List[str] = []
    start = 0
    while start < len(text):
        end = min(start + max_len, len(text))
        piece = text[start:end].strip()
        if piece:
            pieces.append(piece)
        if end >= len(text):
            break
        start = max(end - CHUNK_OVERLAP, start + 1)
    return pieces


def chunk_by_structure(text: str) -> List[ArticleChunk]:
    """
    按监管文件结构切分：章 -> 节 -> 条/点。

    支持两类常见格式:
        1. 办法/准则: 第一章、第一节、第十三条
        2. 指导意见/通知: 二十三、 或 第X点

    若未识别到结构，则整篇作为一条兜底入库。
    """
    if not text:
        return []

    # 匹配章、节、条、点，以及「二十三、」这类枚举条目
    marker_pattern = re.compile(
        rf"(第{CN_NUM}章[^\n]{{0,40}}|"
        rf"第{CN_NUM}节[^\n]{{0,40}}|"
        rf"第{CN_NUM}条|"
        rf"第{CN_NUM}点|"
        rf"{CN_NUM}、)"
    )

    matches = list(marker_pattern.finditer(text))
    if not matches:
        return [ArticleChunk(text=text)]

    current_chapter = ""
    current_section = ""
    current_article = ""
    article_starts: List[Tuple[int, str, str, str]] = []

    for match in matches:
        marker = match.group(0).strip()
        if re.fullmatch(rf"第{CN_NUM}章[^\n]{{0,40}}", marker):
            current_chapter = marker
            current_section = ""
        elif re.fullmatch(rf"第{CN_NUM}节[^\n]{{0,40}}", marker):
            current_section = marker
        elif re.fullmatch(rf"第{CN_NUM}条", marker):
            current_article = marker
            article_starts.append((match.start(), current_chapter, current_section, current_article))
        elif re.fullmatch(rf"第{CN_NUM}点", marker):
            current_article = marker
            article_starts.append((match.start(), current_chapter, current_section, current_article))
        elif re.fullmatch(rf"{CN_NUM}、", marker):
            # 指导意见常见格式: 二十三、
            current_article = f"第{marker.rstrip('、')}点"
            article_starts.append((match.start(), current_chapter, current_section, current_article))

    if not article_starts:
        return [ArticleChunk(text=text)]

    chunks: List[ArticleChunk] = []
    for index, (start_pos, chapter, section, article) in enumerate(article_starts):
        end_pos = article_starts[index + 1][0] if index + 1 < len(article_starts) else len(text)
        body = text[start_pos:end_pos].strip()
        if not body:
            continue

        for piece in split_long_article(body):
            chunks.append(
                ArticleChunk(
                    text=piece,
                    chapter=chapter,
                    section=section,
                    article=article,
                )
            )

    return chunks


def build_chunk_embedding_text(display_name: str, chunk: ArticleChunk) -> str:
    """
    构造用于向量化的文本，显式携带章/节/条信息以提升检索粒度。
    """
    header_parts = [f"【文件】{display_name}"]
    if chunk.chapter:
        header_parts.append(f"【章】{chunk.chapter}")
    if chunk.section:
        header_parts.append(f"【节】{chunk.section}")
    if chunk.article:
        header_parts.append(f"【条】{chunk.article}")
    return "\n".join(header_parts + [chunk.text])


def load_summary_chunks(doc_dir: Path) -> List[TextChunk]:
    """
    读取「监管文件摘要.xlsx」中的结构化摘要。

    摘要表中的「文件名」列已是正确中文，可直接作为高质量检索片段。
    """
    summary_path = find_xlsx_by_header(doc_dir, SUMMARY_XLSX_HEADER)
    if summary_path is None:
        return []

    workbook = openpyxl.load_workbook(summary_path, read_only=True, data_only=True)
    worksheet = workbook.active
    chunks: List[TextChunk] = []
    chunk_id = 0
    current_display_name = ""

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        file_name, chapter, summary, detail = row[0], row[1], row[2], row[3]
        if file_name:
            current_display_name = str(file_name).strip()

        detail_text = normalize_text(str(detail or ""))
        if not detail_text:
            continue

        chapter_text = str(chapter or "").strip()
        summary_text = str(summary or "").strip()

        # 尝试从摘要表的章节字段解析「章/节/条」
        parsed_chapter, parsed_section, parsed_article = parse_chapter_label(chapter_text)
        merged_text = build_chunk_embedding_text(
            current_display_name or "监管文件摘要",
            ArticleChunk(
                text=detail_text,
                chapter=parsed_chapter,
                section=parsed_section,
                article=parsed_article,
            ),
        )
        if summary_text:
            merged_text = f"【概述】{summary_text}\n{merged_text}"

        chunks.append(
            TextChunk(
                chunk_id=chunk_id,
                text=merged_text,
                source_type="summary_xlsx",
                raw_filename=summary_path.name,
                display_name=current_display_name or "监管文件摘要",
                chapter=parsed_chapter or chapter_text,
                section=parsed_section,
                article=parsed_article,
                summary=summary_text,
            )
        )
        chunk_id += 1

    workbook.close()
    return chunks


def parse_chapter_label(label: str) -> Tuple[str, str, str]:
    """
    从摘要表或目录中的章节描述解析结构信息。

    示例:
        「第三章-第一节-第十三条（P4）」 -> 第三章, 第一节, 第十三条
        「第二十三点（P16）」           -> "", "", 第二十三点
    """
    if not label:
        return "", "", ""

    chapter_match = re.search(rf"第{CN_NUM}章[^\-]*", label)
    section_match = re.search(rf"第{CN_NUM}节[^\-]*", label)
    article_match = re.search(rf"第{CN_NUM}[条点][^\（\(]*", label)

    chapter = chapter_match.group(0).strip() if chapter_match else ""
    section = section_match.group(0).strip() if section_match else ""
    article = article_match.group(0).strip() if article_match else ""
    return chapter, section, article


def load_document_chunks(doc_dir: Path, catalog: Sequence[CatalogEntry]) -> List[TextChunk]:
    """读取 PDF / DOC 文件，按章/节/条切分为 TextChunk。"""
    chunks: List[TextChunk] = []
    chunk_id = 0

    doc_files = sorted(
        path
        for path in doc_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".pdf", ".doc", ".docx"}
    )

    for file_path in doc_files:
        print(f"  解析文档: {file_path.name}")
        try:
            raw_text = extract_document_text(file_path)
        except Exception as exc:  # noqa: BLE001 - 单文件失败不应中断全流程
            print(f"    [跳过] 解析失败: {exc}")
            continue

        raw_text = normalize_text(raw_text)
        if not raw_text:
            print("    [跳过] 未提取到有效文本")
            continue

        catalog_entry = match_file_to_catalog(file_path.name, raw_text, catalog)
        display_name = catalog_entry.display_name
        structured_chunks = chunk_by_structure(raw_text)
        print(f"    -> 映射名称: {display_name} | 切分 {len(structured_chunks)} 条")

        for piece in structured_chunks:
            chunks.append(
                TextChunk(
                    chunk_id=chunk_id,
                    text=build_chunk_embedding_text(display_name, piece),
                    source_type="document",
                    raw_filename=file_path.name,
                    display_name=display_name,
                    chapter=piece.chapter,
                    section=piece.section,
                    article=piece.article,
                )
            )
            chunk_id += 1

    return chunks


def collect_all_chunks(doc_dir: Path) -> List[TextChunk]:
    """汇总文档块与摘要块，并重新编号。"""
    catalog = load_catalog(doc_dir)
    doc_chunks = load_document_chunks(doc_dir, catalog)
    summary_chunks = load_summary_chunks(doc_dir)

    all_chunks = doc_chunks + summary_chunks
    for index, chunk in enumerate(all_chunks):
        chunk.chunk_id = index
    return all_chunks


# =============================================================================
# Embedding 与 FAISS
# =============================================================================


def embed_texts(client: OpenAI, texts: Sequence[str]) -> np.ndarray:
    """
    调用百炼 Qwen3-Embedding 生成向量。

    返回:
        shape = (n, EMBEDDING_DIM) 的 float32 数组
    """
    if not texts:
        return np.zeros((0, EMBEDDING_DIM), dtype="float32")

    all_vectors: List[List[float]] = []
    total = len(texts)

    for start in range(0, total, EMBED_BATCH_SIZE):
        batch = list(texts[start : start + EMBED_BATCH_SIZE])
        response = client.embeddings.create(
            model=EMBEDDING_MODEL,
            input=batch,
            dimensions=EMBEDDING_DIM,
            encoding_format="float",
        )
        batch_vectors = [item.embedding for item in response.data]
        all_vectors.extend(batch_vectors)
        print(f"    已向量化 {min(start + len(batch), total)}/{total} 条")

    vectors = np.array(all_vectors, dtype="float32")
    return vectors


def build_faiss_index(vectors: np.ndarray) -> faiss.Index:
    """
    构建 FAISS 索引。

    说明:
        向量先做 L2 归一化，再使用 IndexFlatIP，等价于余弦相似度检索。
    """
    if vectors.size == 0:
        raise ValueError("向量数组为空，无法构建 FAISS 索引。")

    faiss.normalize_L2(vectors)
    index = faiss.IndexFlatIP(vectors.shape[1])
    index.add(vectors)
    return index


def save_index_bundle(
    index: faiss.Index,
    chunks: Sequence[TextChunk],
    vectors: np.ndarray,
) -> None:
    """持久化 FAISS 索引、元数据与配置。"""
    INDEX_DIR.mkdir(parents=True, exist_ok=True)
    faiss.write_index(index, str(FAISS_INDEX_FILE))

    metadata = [asdict(chunk) for chunk in chunks]
    with METADATA_FILE.open("wb") as file:
        pickle.dump(metadata, file)

    config = {
        "embedding_model": EMBEDDING_MODEL,
        "embedding_dim": EMBEDDING_DIM,
        "chunk_strategy": "chapter_section_article",
        "max_article_len": MAX_ARTICLE_LEN,
        "chunk_overlap": CHUNK_OVERLAP,
        "vector_count": len(chunks),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    CONFIG_FILE.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n[完成] FAISS 索引: {FAISS_INDEX_FILE}")
    print(f"[完成] 元数据文件: {METADATA_FILE}")
    print(f"[完成] 向量条目数: {vectors.shape[0]}")


def load_index_bundle() -> Tuple[faiss.Index, List[dict], dict]:
    """加载本地 FAISS 索引与元数据。"""
    if not FAISS_INDEX_FILE.exists() or not METADATA_FILE.exists():
        raise FileNotFoundError(
            f"未找到索引文件，请先运行: python {Path(__file__).name}"
        )

    index = faiss.read_index(str(FAISS_INDEX_FILE))
    with METADATA_FILE.open("rb") as file:
        metadata = pickle.load(file)

    config = {}
    if CONFIG_FILE.exists():
        config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))

    return index, metadata, config


def search(
    client: OpenAI,
    query: str,
    top_k: int = 5,
) -> List[Tuple[float, dict]]:
    """
    语义检索入口。

    返回:
        [(相似度, 元数据), ...]
    """
    index, metadata, _ = load_index_bundle()
    query_vector = embed_texts(client, [query])
    faiss.normalize_L2(query_vector)

    scores, indices = index.search(query_vector, top_k)
    results: List[Tuple[float, dict]] = []

    for score, idx in zip(scores[0], indices[0]):
        if idx < 0:
            continue
        results.append((float(score), metadata[idx]))

    return results


def build_index(force: bool = False) -> None:
    """执行完整的入库流程。"""
    if FAISS_INDEX_FILE.exists() and not force:
        print("检测到已有索引。若需重建，请加参数 --rebuild")
        return

    if not DOC_DIR.exists():
        raise FileNotFoundError(f"监管文件目录不存在: {DOC_DIR}")

    print("=" * 60)
    print("  监管文件 Embedding 入库")
    print("=" * 60)

    api_key = load_api_key()
    client = create_embedding_client(api_key)

    print("\n[1/4] 读取并切分文档...")
    chunks = collect_all_chunks(DOC_DIR)
    if not chunks:
        raise RuntimeError("没有可用于向量化的文本块。")
    print(f"      共生成 {len(chunks)} 个文本块")

    print("\n[2/4] 调用百炼 Qwen3-Embedding 生成向量...")
    texts = [chunk.text for chunk in chunks]
    vectors = embed_texts(client, texts)

    print("\n[3/4] 构建 FAISS 索引...")
    index = build_faiss_index(vectors)

    print("\n[4/4] 保存索引与元数据...")
    save_index_bundle(index, chunks, vectors)


def format_location(item: dict) -> str:
    """拼接章/节/条定位信息。"""
    parts = []
    if item.get("chapter"):
        parts.append(item["chapter"])
    if item.get("section"):
        parts.append(item["section"])
    if item.get("article"):
        parts.append(item["article"])
    return " / ".join(parts)


def print_search_results(results: Sequence[Tuple[float, dict]]) -> None:
    """格式化输出检索结果。"""
    if not results:
        print("未检索到结果。")
        return

    for rank, (score, item) in enumerate(results, start=1):
        print("\n" + "-" * 60)
        print(f"Top {rank} | 相似度: {score:.4f}")
        print(f"文件: {item.get('display_name', '')}")
        location = format_location(item)
        if location:
            print(f"定位: {location}")
        if item.get("summary"):
            print(f"概述: {item['summary']}")
        preview = item.get("text", "").replace("\n", " ")
        print(f"内容: {preview[:220]}{'...' if len(preview) > 220 else ''}")


# =============================================================================
# 命令行入口
# =============================================================================


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""
    parser = argparse.ArgumentParser(description="监管文件 Embedding + FAISS 入库工具")
    parser.add_argument(
        "--search",
        type=str,
        default="",
        help="对已建索引执行语义检索",
    )
    parser.add_argument(
        "--top-k",
        type=int,
        default=5,
        help="检索返回条数（默认 5）",
    )
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="强制重建 FAISS 索引",
    )
    return parser.parse_args()


def main() -> None:
    """脚本主入口。"""
    args = parse_args()

    try:
        if args.search:
            api_key = load_api_key()
            client = create_embedding_client(api_key)
            print(f"检索问题: {args.search}")
            results = search(client, args.search, top_k=args.top_k)
            print_search_results(results)
            return

        build_index(force=args.rebuild)
    except Exception as exc:  # noqa: BLE001 - 统一输出错误信息
        print(f"\n[错误] {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
